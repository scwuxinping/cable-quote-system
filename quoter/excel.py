"""Excel 导入询价 / 导出报价单 / 下载导入模板。"""
import io
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import Quotation
from .pricing import resolve_series_layout

THIN = Side(style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill('solid', fgColor='DDEBF7')


def _cell(ws, row, col, value, *, bold=False, fill=None, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    return c


def empty_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '询价清单'
    headers = ['型号', '规格', '长度(m)']
    for i, h in enumerate(headers, 1):
        _cell(ws, 1, i, h, bold=True, fill=HEAD_FILL)
    examples = [('YJV', '3x120+1x70', 500), ('BV', '1x2.5', 1000), ('VV', '4x16', 300)]
    for r, row in enumerate(examples, 2):
        for c, v in enumerate(row, 1):
            _cell(ws, r, c, v)
    for col, width in zip('ABC', (16, 20, 12)):
        ws.column_dimensions[col].width = width
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=import_template.xlsx'
    wb.save(resp)
    return resp


def parse_inquiry(file_obj):
    """解析上传的询价 Excel。返回 (items, errors)。

    items: [(spec, Decimal 长度)]；errors: [(行号, 原文, 原因)]
    """
    from .models import CableSeries, CableSpec

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    items = []
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not row or all(v in (None, '') for v in row):
            continue
        model, spec_text, length = (list(row) + [None, None, None])[:3]
        if str(model or '').strip() in ('型号',) and str(spec_text or '').strip() in ('规格', '规格型号'):
            continue  # 表头
        raw = '%s %s' % (model or '', spec_text or '')
        try:
            length = Decimal(str(length).strip())
        except (InvalidOperation, AttributeError, ValueError):
            errors.append((idx, raw.strip(), '长度不是数字'))
            continue
        if length <= 0:
            errors.append((idx, raw.strip(), '长度必须大于 0'))
            continue
        try:
            series, layout = resolve_series_layout(raw)
        except ValueError as exc:
            errors.append((idx, raw.strip(), str(exc)))
            continue
        if series is None:
            errors.append((idx, raw.strip(), '型号无法识别'))
            continue
        spec = series.specs.filter(layout=layout).first()
        if spec is None:
            errors.append((idx, raw.strip(), '规格库中无此规格，可先到规格库添加'))
            continue
        items.append((spec, length))
    return items, errors


def export_quotation_bytes(quotation):
    """生成报价单 Excel 的字节流（邮件附件与下载共用）。"""
    from .models import PricingParams
    p = PricingParams.load()
    wb = Workbook()
    ws = wb.active
    ws.title = '报价单'
    for col, width in zip('ABCDEFGH', (6, 30, 12, 10, 14, 12, 14, 16)):
        ws.column_dimensions[col].width = width

    _cell(ws, 1, 1, p.company_name, bold=True, align='left')
    ws.merge_cells('A1:H1')
    _cell(ws, 2, 1, '报 价 单', bold=True)
    ws.merge_cells('A2:H2')
    _cell(ws, 3, 1, '单号：%s' % quotation.number, align='left')
    ws.merge_cells('A3:D3')
    _cell(ws, 3, 5, '日期：%s' % quotation.created_at.strftime('%Y-%m-%d'), align='left')
    ws.merge_cells('E3:H3')
    _cell(ws, 4, 1, '客户：%s' % quotation.customer.name, align='left')
    ws.merge_cells('A4:D4')
    _cell(ws, 4, 5, '有效期至：%s' % quotation.valid_until, align='left')
    ws.merge_cells('E4:H4')
    _cell(ws, 5, 1, '铜价基准：%s 元/kg%s' % (
        quotation.base_cu_price,
        '（%s）' % quotation.cu_price_source if quotation.cu_price_source else ''), align='left')
    ws.merge_cells('A5:D5')
    _cell(ws, 5, 5, '报价口径：%s' % ('含税' if p.price_with_tax else '不含税'), align='left')
    ws.merge_cells('E5:H5')

    headers = ['序号', '型号规格', '电压', '单位', '单价(元/m)', '长度(m)', '金额(元)', '备注']
    for c, h in enumerate(headers, 1):
        _cell(ws, 7, c, h, bold=True, fill=HEAD_FILL)

    r = 8
    for i, item in enumerate(quotation.items.all(), 1):
        _cell(ws, r, 1, i)
        _cell(ws, r, 2, item.spec_text or str(item.spec), align='left')
        _cell(ws, r, 3, item.spec.voltage)
        _cell(ws, r, 4, '米')
        _cell(ws, r, 5, float(item.final_price_per_m))
        _cell(ws, r, 6, float(item.length_m))
        _cell(ws, r, 7, float(item.amount))
        _cell(ws, r, 8, '')
        r += 1

    goods_total = sum((item.amount for item in quotation.items.all()), Decimal('0'))
    _cell(ws, r, 6, '货款小计', bold=True)
    _cell(ws, r, 7, float(goods_total), bold=True)
    r += 1
    _cell(ws, r, 6, '运费', bold=True)
    _cell(ws, r, 7, float(quotation.freight or Decimal('0')), bold=True)
    r += 1
    _cell(ws, r, 6, '合计', bold=True)
    _cell(ws, r, 7, float(quotation.total_amount()), bold=True)
    r += 2
    _cell(ws, r, 1, p.quote_note, align='left')
    ws.merge_cells('A%d:H%d' % (r, r))
    if p.company_phone:
        r += 1
        _cell(ws, r, 1, '联系电话：%s' % p.company_phone, align='left')
        ws.merge_cells('A%d:H%d' % (r, r))
    if p.company_address:
        r += 1
        _cell(ws, r, 1, '地址：%s' % p.company_address, align='left')
        ws.merge_cells('A%d:H%d' % (r, r))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_quotation(quotation):
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=%s.xlsx' % quotation.number
    resp.write(export_quotation_bytes(quotation))
    return resp
